"""CSV / Excel export.

Python port of the original C# Forms/Export.xaml.cs + the Models/Export/* maps.

Collects the four data sets (monthly, daily, half-hourly, meter readings) out of
the account's SQLite database, merging electric + gas onto shared date keys, and
renders them as either:

  * a ZIP of four CSV files, or
  * a single .xlsx workbook with five sheets (adds a pivoted "Monthly Chart").

Column headers and number formats mirror the original spreadsheet maps. Unlike
the original, CSV uses the same friendly headers as Excel (the C# CSV path used
raw property names); consistent headers are the more useful behaviour.
"""
from __future__ import annotations

import csv
import datetime as dt
import io
import zipfile

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .database import Database

# Excel number formats, from the original Models/Export/Styles.cs
FMT_KWH = '0.00"kwh"'
FMT_POUNDS = '£#,##0.00'
FMT_PENCE = '0.0000"p"'
FMT_GENERAL = '#,##0.0##,'
FMT_TIMESTAMP = 'yyyy-mm-dd hh:mm'

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _proper_case(text: str) -> str:
    return text[0].upper() + text[1:].lower() if text else text


# ---------------------------------------------------------------- collection
def collect_monthly(db: Database) -> list[dict]:
    merged: dict[str, dict] = {}
    for row in db.fetch_monthly("Electric"):
        merged[row["Month"]] = {
            "month": row["Month"],
            "electricKwh": row["Consumption"], "electricCost": row["Cost"],
            "gasKwh": 0.0, "gasCost": 0.0,
        }
    for row in db.fetch_monthly("Gas"):
        m = merged.setdefault(row["Month"], {
            "month": row["Month"], "electricKwh": 0.0, "electricCost": 0.0,
            "gasKwh": 0.0, "gasCost": 0.0})
        m["gasKwh"] = row["Consumption"]
        m["gasCost"] = row["Cost"]
    return [merged[k] for k in sorted(merged, reverse=True)]


def collect_daily(db: Database) -> list[dict]:
    merged: dict[str, dict] = {}
    for row in db.fetch_daily("Electric"):
        merged[row["Day"]] = {
            "day": row["Day"],
            "electricKwh": row["Consumption"], "electricStanding": row["Standing"],
            "electricAnyTime": row["AnyTime"], "electricPeak": row["Peak"],
            "electricOffPeak": row["OffPeak"], "electricCost": row["Cost"],
            "gasKwh": 0.0, "gasStanding": 0.0, "gasAnyTime": 0.0, "gasCost": 0.0,
        }
    for row in db.fetch_daily("Gas"):
        m = merged.setdefault(row["Day"], {
            "day": row["Day"], "electricKwh": 0.0, "electricStanding": 0.0,
            "electricAnyTime": 0.0, "electricPeak": 0.0, "electricOffPeak": 0.0,
            "electricCost": 0.0, "gasKwh": 0.0, "gasStanding": 0.0,
            "gasAnyTime": 0.0, "gasCost": 0.0})
        m["gasKwh"] = row["Consumption"]
        m["gasStanding"] = row["Standing"]
        m["gasAnyTime"] = row["AnyTime"]
        m["gasCost"] = row["Cost"]
    return [merged[k] for k in sorted(merged, reverse=True)]


def collect_half_hourly(db: Database) -> list[dict]:
    merged: dict[str, dict] = {}
    for row in db.fetch_half_hourly("Electric"):
        merged[row["StartTime"]] = {
            "startTime": row["StartTime"], "electricKwh": row["Consumption"], "gasKwh": 0.0}
    for row in db.fetch_half_hourly("Gas"):
        m = merged.setdefault(row["StartTime"], {
            "startTime": row["StartTime"], "electricKwh": 0.0, "gasKwh": 0.0})
        m["gasKwh"] = row["Consumption"]
    return [merged[k] for k in sorted(merged, reverse=True)]


def _parse_date(value: str, default: dt.date) -> dt.date:
    try:
        return dt.date.fromisoformat(value[:10])
    except (ValueError, TypeError):
        return default


def collect_readings(db: Database) -> list[dict]:
    # Build the register list so each reading can be tagged with its unit.
    registers = []
    for reg in db.fetch_meter_registers():
        registers.append({
            "start": _parse_date(reg["StartDate"], dt.date.min),
            "end": _parse_date(reg["EndDate"], dt.date.max) if reg["EndDate"] else dt.date.max,
            "fuelType": reg["FuelType"],
            "unit": reg["UnitOfMeasurement"],
        })

    result: dict[str, dict] = {}
    for reading in db.fetch_meter_readings():
        rdate = _parse_date(reading["Date"], dt.date.min)
        unit = next(
            (r["unit"] for r in registers
             if r["fuelType"] == reading["FuelType"] and r["start"] < rdate <= r["end"]),
            "",
        )
        key = f"{reading['Date']}-{reading['FuelType']}-{reading['TimingCategory']}"
        result[key] = {
            "date": reading["Date"],
            "fuelType": _proper_case(reading["FuelType"]),
            "category": reading["TimingCategory"],
            "value": reading["Value"],
            "unit": unit,
        }
    return [result[k] for k in sorted(result)]


def collect_monthly_chart(monthly: list[dict]) -> list[dict]:
    by_year: dict[str, dict] = {}
    for row in monthly:
        year, month = row["month"].split("-")
        idx = int(month) - 1
        entry = by_year.setdefault(year, {"year": year,
                                          "kwh": [0.0] * 12, "cost": [0.0] * 12})
        entry["kwh"][idx] = row["electricKwh"] + row["gasKwh"]
        entry["cost"][idx] = row["electricCost"] + row["gasCost"]
    return [by_year[y] for y in sorted(by_year)]


# ---------------------------------------------------------------------- CSV
_CSV_SPECS = {
    "Monthly": (
        ["Date", "Electric kWh", "Electric Cost", "Gas kWh", "Gas Cost"],
        lambda r: [r["month"], r["electricKwh"], r["electricCost"], r["gasKwh"], r["gasCost"]],
    ),
    "Daily": (
        ["Day", "Electric kWh", "Electric Standing", "Electric AnyTime", "Electric Peak",
         "Electric Off Peak", "Electric Cost", "Gas kWh", "Gas Standing", "Gas AnyTime", "Gas Cost"],
        lambda r: [r["day"], r["electricKwh"], r["electricStanding"], r["electricAnyTime"],
                   r["electricPeak"], r["electricOffPeak"], r["electricCost"],
                   r["gasKwh"], r["gasStanding"], r["gasAnyTime"], r["gasCost"]],
    ),
    "Half Hourly": (
        ["Start Time", "Electric kWh", "Gas kWh"],
        lambda r: [r["startTime"], r["electricKwh"], r["gasKwh"]],
    ),
    "Meter Readings": (
        ["Date", "Fuel Type", "Category", "Reading", "Units"],
        lambda r: [r["date"], r["fuelType"], r["category"], r["value"], r["unit"]],
    ),
}


def build_csv_zip(account_id: str) -> bytes:
    db = Database(account_id)
    data = {
        "Monthly": collect_monthly(db),
        "Daily": collect_daily(db),
        "Half Hourly": collect_half_hourly(db),
        "Meter Readings": collect_readings(db),
    }
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, rows in data.items():
            headers, row_fn = _CSV_SPECS[name]
            text = io.StringIO()
            writer = csv.writer(text)
            writer.writerow(headers)
            for r in rows:
                writer.writerow(row_fn(r))
            zf.writestr(f"{account_id} {name}.csv", text.getvalue())
    return buffer.getvalue()


# -------------------------------------------------------------------- Excel
def _write_sheet(ws, headers, rows, formats=None):
    ws.append(headers)
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    if formats:
        # formats: {col_index (0-based): number_format}
        for col_idx, fmt in formats.items():
            letter = get_column_letter(col_idx + 1)
            for cell in ws[letter][1:]:  # skip header
                cell.number_format = fmt
    # Approximate auto-fit.
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(str(header)) + 2)


def build_excel(account_id: str) -> bytes:
    db = Database(account_id)
    monthly = collect_monthly(db)
    daily = collect_daily(db)
    half_hourly = collect_half_hourly(db)
    readings = collect_readings(db)
    chart = collect_monthly_chart(monthly)

    wb = Workbook()

    # Monthly
    ws = wb.active
    ws.title = "Monthly"
    _write_sheet(
        ws,
        ["Date", "Electric kWh", "Electric Cost", "Gas kWh", "Gas Cost"],
        [[r["month"], r["electricKwh"], r["electricCost"], r["gasKwh"], r["gasCost"]] for r in monthly],
        {1: FMT_KWH, 2: FMT_POUNDS, 3: FMT_KWH, 4: FMT_POUNDS},
    )

    # Monthly Chart (pivoted by year)
    ws = wb.create_sheet("Monthly Chart")
    chart_headers = ["Year"] + [f"{m} kWh" for m in MONTHS] + [f"{m} Cost" for m in MONTHS]
    chart_rows = [[c["year"], *c["kwh"], *c["cost"]] for c in chart]
    chart_formats = {i: FMT_KWH for i in range(1, 13)}
    chart_formats.update({i: FMT_POUNDS for i in range(13, 25)})
    _write_sheet(ws, chart_headers, chart_rows, chart_formats)

    # Daily
    ws = wb.create_sheet("Daily")
    _write_sheet(
        ws,
        ["Day", "Electric kWh", "Electric Standing", "Electric AnyTime", "Electric Peak",
         "Electric Off Peak", "Electric Cost", "Gas kWh", "Gas Standing", "Gas AnyTime", "Gas Cost"],
        [[r["day"], r["electricKwh"], r["electricStanding"], r["electricAnyTime"], r["electricPeak"],
          r["electricOffPeak"], r["electricCost"], r["gasKwh"], r["gasStanding"], r["gasAnyTime"],
          r["gasCost"]] for r in daily],
        {1: FMT_KWH, 2: FMT_PENCE, 3: FMT_PENCE, 4: FMT_PENCE, 5: FMT_PENCE, 6: FMT_POUNDS,
         7: FMT_KWH, 8: FMT_PENCE, 9: FMT_PENCE, 10: FMT_POUNDS},
    )

    # Half Hourly (Start Time as real datetime)
    ws = wb.create_sheet("Half Hourly")
    hh_rows = []
    for r in half_hourly:
        try:
            ts = dt.datetime.strptime(r["startTime"], "%Y-%m-%d %H:%M:%S")
        except (ValueError, TypeError):
            ts = r["startTime"]
        hh_rows.append([ts, r["electricKwh"], r["gasKwh"]])
    _write_sheet(ws, ["Start Time", "Electric kWh", "Gas kWh"], hh_rows,
                 {0: FMT_TIMESTAMP, 1: FMT_KWH, 2: FMT_KWH})
    ws.column_dimensions["A"].width = 20

    # Meter Readings
    ws = wb.create_sheet("Meter Readings")
    _write_sheet(
        ws,
        ["Date", "Fuel Type", "Category", "Reading", "Units"],
        [[r["date"], r["fuelType"], r["category"], r["value"], r["unit"]] for r in readings],
        {3: FMT_GENERAL},
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
